# app.py
import io
import os
import re
import fitz  # PyMuPDF
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from openai import OpenAI

# ==========================================
# 1. DATA MODELS & SCHEMAS
# ==========================================

class IronmongeryItem(BaseModel):
    item_type: str = Field(description="Closer, Lockset, Panic Bar, Handle, Vision Panel, etc.")
    description: str

class DoorItem(BaseModel):
    door_ref: str = Field(default="D-01", description="Door tag or mark")
    quantity: int = Field(default=1, description="Quantity of door sets")
    width_mm: Optional[int] = Field(default=1000, description="Width in mm (Structural/Frame)")
    height_mm: Optional[int] = Field(default=2100, description="Height in mm (Structural/Frame)")
    dimension_basis: str = Field(default="Structural Opening", description="Structural Opening, Over-Frame, or Clear Opening")
    configuration: str = Field(default="Single", description="Single, Double, Leaf-and-Half")
    fire_rating: str = Field(default="Unrated", description="Unrated, FD30, FD60, FD120, etc.")
    security_rating: str = Field(default="Standard", description="Standard, PAS 24, SR2, SR3, RC3")
    finish: str = Field(default="PPC RAL 7016 Matt", description="Powder coat color or primed")
    ironmongery_summary: str = Field(default="Sashlock, Lever Handles, Overhead Closer", description="Formatted hardware summary")
    local_notes: str = Field(default="", description="Handing, vision panel size, threshold detail")

class BatchExtractionResult(BaseModel):
    project_name: Optional[str] = Field(None, description="Project Name")
    project_reference: Optional[str] = Field(None, description="Drawing or Enquiry Reference")
    revision: Optional[str] = Field("A", description="Revision code")
    doors: List[DoorItem] = []
    global_notes: List[str] = []


# ==========================================
# 2. EXTRACTION ENGINE
# ==========================================

def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Extracts raw text and spatial layout strings from an uploaded PDF."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages_text = []
    for i, page in enumerate(doc):
        text = page.get_text("text").strip()
        if text:
            pages_text.append(f"--- PAGE {i + 1} ---\n{text}")
    doc.close()
    return "\n\n".join(pages_text)

def parse_schedule_with_llm(raw_text: str, api_key: str) -> BatchExtractionResult:
    """Parses raw text into structured door schedule data using OpenAI Structured Outputs."""
    client = OpenAI(api_key=api_key)
    prompt = (
        "You are an expert steel hinged door technical estimator. Extract all door schedule items, "
        "dimensions, performance ratings, and ironmongery requirements from the text. "
        "Standardise dimensions into integer millimeters (mm). "
        "Separate schedule-wide assumptions into global_notes, and item-specific details "
        "(handing, vision panel sizes, drop seals) into local_notes for the relevant door."
    )
    
    response = client.beta.chat.completions.parse(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": prompt},
            {"role": "user", "content": raw_text}
        ],
        response_format=BatchExtractionResult,
        temperature=0.0
    )
    return response.choices[0].message.parsed


# ==========================================
# 3. EXCEL GENERATION ENGINE (IN-MEMORY)
# ==========================================

def generate_excel_quote_bytes(
    metadata: Dict[str, str], 
    df_doors: pd.DataFrame, 
    global_notes: List[str]
) -> bytes:
    """Creates a formatted Excel workbook in-memory and returns its byte stream."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Door Takeoff"

    # Styling Palettes
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    meta_font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    border_thin = Side(border_style="thin", color="D9D9D9")
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # 1. Project Metadata Block
    ws["A1"] = "STEEL HINGED DOOR QUOTATION SCHEDULE"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    
    ws["A3"] = "Project Name:"
    ws["B3"] = metadata.get("Project", "N/A")
    ws["A4"] = "Enquiry Ref:"
    ws["B4"] = metadata.get("Reference", "N/A")
    ws["D3"] = "Revision:"
    ws["E3"] = metadata.get("Revision", "A")
    ws["D4"] = "Date:"
    ws["E4"] = metadata.get("Date", "17/08/2026")

    for cell_ref in ["A3", "A4", "D3", "D4"]:
        ws[cell_ref].font = meta_font_bold

    # 2. Table Column Headers
    headers = [
        "Mark", "Qty", "Width (mm)", "Height (mm)", "Basis", 
        "Config", "Fire Rating", "Security", "Finish", 
        "Ironmongery / Hardware Summary", "Notes & Handing"
    ]
    
    start_row = 7
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3. Populate Door Records
    current_row = start_row + 1
    for _, row in df_doors.iterrows():
        ws.cell(row=current_row, column=1, value=str(row.get("door_ref", "")))
        ws.cell(row=current_row, column=2, value=int(row.get("quantity", 1)))
        ws.cell(row=current_row, column=3, value=row.get("width_mm"))
        ws.cell(row=current_row, column=4, value=row.get("height_mm"))
        ws.cell(row=current_row, column=5, value=str(row.get("dimension_basis", "")))
        ws.cell(row=current_row, column=6, value=str(row.get("configuration", "")))
        ws.cell(row=current_row, column=7, value=str(row.get("fire_rating", "")))
        ws.cell(row=current_row, column=8, value=str(row.get("security_rating", "")))
        ws.cell(row=current_row, column=9, value=str(row.get("finish", "")))
        ws.cell(row=current_row, column=10, value=str(row.get("ironmongery_summary", "")))
        ws.cell(row=current_row, column=11, value=str(row.get("local_notes", "")))

        # Format row borders and alignment
        for c in range(1, 12):
            cell = ws.cell(row=current_row, column=c)
            cell.border = box_border
            if c in [1, 2, 3, 4, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        current_row += 1

    # 4. Global Notes & Clarifications Footer
    notes_start = current_row + 2
    ws.cell(row=notes_start, column=1, value="GLOBAL SCHEDULE NOTES & CLARIFICATIONS:").font = meta_font_bold
    
    for idx, note in enumerate(global_notes, start=1):
        ws.cell(row=notes_start + idx, column=1, value=f"• {note}")

    # Column Width Auto-Fitting
    col_widths = {1: 12, 2: 8, 3: 14, 4: 14, 5: 18, 6: 14, 7: 14, 8: 14, 9: 22, 10: 45, 11: 35}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Save to memory buffer
    output_stream = io.BytesIO()
    wb.save(output_stream)
    return output_stream.getvalue()


# ==========================================
# 4. STREAMLIT USER INTERFACE
# ==========================================

st.set_page_config(
    page_title="Steel Door Quoting Engine",
    page_icon="🚪",
    layout="wide"
)

st.title("🚪 Steel Hinged Door Estimating & Takeoff Engine")
st.markdown("Upload inbound enquiry emails or PDF drawing schedules, review extracted parameters, and export to Excel.")

# Sidebar Configuration
with st.sidebar:
    st.header("⚙️ Configuration")
    api_key_input = st.text_input("OpenAI API Key", type="password", value=os.getenv("OPENAI_API_KEY", ""))
    st.markdown("---")
    st.markdown("### Default Assumptions")
    st.caption("• Standard steel gauge: 1.5mm leaf / 2.0mm frame\n• Dimensions defaulted to Structural Opening\n• Unrated doors priced as standard primed/PPC")

# Initialize Session State
if "extracted_data" not in st.session_state:
    st.session_state["extracted_data"] = None

# Input Section: PDF or Text Area
tab1, tab2 = st.tabs(["📄 Upload PDF Schedule", "✉️ Paste Email / Text Enquiry"])

with tab1:
    uploaded_file = st.file_uploader("Upload Architectural Schedule (PDF)", type=["pdf"])
    if uploaded_file and st.button("Parse PDF Document", type="primary"):
        if not api_key_input:
            st.error("Please provide an OpenAI API Key in the sidebar.")
        else:
            with st.spinner("Extracting layout grids and analyzing schedule..."):
                raw_text = extract_text_from_pdf(uploaded_file.getvalue())
                result = parse_schedule_with_llm(raw_text, api_key_input)
                st.session_state["extracted_data"] = result
                st.success("Extraction complete!")

with tab2:
    email_text = st.text_area("Paste Raw Email Body / Schedule Text", height=200, placeholder="Paste RFQ email text here...")
    if st.button("Parse Text Enquiry", type="primary"):
        if not api_key_input:
            st.error("Please provide an OpenAI API Key in the sidebar.")
        elif not email_text.strip():
            st.warning("Please paste enquiry text first.")
        else:
            with st.spinner("Processing door parameters..."):
                result = parse_schedule_with_llm(email_text, api_key_input)
                st.session_state["extracted_data"] = result
                st.success("Extraction complete!")

# Data Review & Interactive Editor
if st.session_state["extracted_data"]:
    data: BatchExtractionResult = st.session_state["extracted_data"]
    
    st.markdown("---")
    st.header("📋 Review & Edit Extracted Schedule")
    
    # Metadata Row
    m_col1, m_col2, m_col3, m_col4 = st.columns(4)
    with m_col1:
        proj_name = st.text_input("Project Name", value=data.project_name or "Industrial Facility")
    with m_col2:
        proj_ref = st.text_input("Drawing / Quote Ref", value=data.project_reference or "RFQ-26-001")
    with m_col3:
        proj_rev = st.text_input("Revision", value=data.revision or "A")
    with m_col4:
        proj_date = st.text_input("Date", value="17/08/2026")

    meta_dict = {
        "Project": proj_name,
        "Reference": proj_ref,
        "Revision": proj_rev,
        "Date": proj_date
    }

    # Convert doors to editable DataFrame
    doors_dicts = [d.model_dump() for d in data.doors]
    df_raw = pd.DataFrame(doors_dicts) if doors_dicts else pd.DataFrame(columns=list(DoorItem.model_fields.keys()))
    
    st.subheader("Door Sets Takeoff Matrix")
    st.caption("You can edit values, add new rows, or delete existing rows directly in the table below.")

    edited_df = st.data_editor(
        df_raw,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "door_ref": st.column_config.TextColumn("Mark / Tag", required=True),
            "quantity": st.column_config.NumberColumn("Qty", min_value=1, step=1, required=True),
            "width_mm": st.column_config.NumberColumn("Width (mm)", min_value=500, max_value=5000, step=5),
            "height_mm": st.column_config.NumberColumn("Height (mm)", min_value=1000, max_value=5000, step=5),
            "dimension_basis": st.column_config.SelectboxColumn("Basis", options=["Structural Opening", "Over-Frame", "Clear Opening"]),
            "configuration": st.column_config.SelectboxColumn("Config", options=["Single", "Double", "Leaf-and-Half"]),
            "fire_rating": st.column_config.SelectboxColumn("Fire Rating", options=["Unrated", "FD30", "FD60", "FD120", "FD240"]),
            "security_rating": st.column_config.SelectboxColumn("Security", options=["Standard", "PAS 24", "LPS 1175 SR2", "LPS 1175 SR3", "RC3"]),
            "finish": st.column_config.TextColumn("Finish"),
            "ironmongery_summary": st.column_config.TextColumn("Hardware Summary", width="large"),
            "local_notes": st.column_config.TextColumn("Local Notes & Handing", width="medium"),
        }
    )

    # Global Notes & Clarifications Editor
    st.subheader("Global Notes & Technical Clarifications")
    notes_text = "\n".join(data.global_notes)
    edited_notes_text = st.text_area("Global Notes (one per line)", value=notes_text, height=120)
    final_global_notes = [n.strip() for n in edited_notes_text.split("\n") if n.strip()]

    # Excel Generation & Download
    st.markdown("---")
    excel_bytes = generate_excel_quote_bytes(meta_dict, edited_df, final_global_notes)
    
    file_name = f"Quote_{proj_ref.replace('/', '_')}.xlsx"
    st.download_button(
        label="📥 Download Formatted Excel Quotation (.xlsx)",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
